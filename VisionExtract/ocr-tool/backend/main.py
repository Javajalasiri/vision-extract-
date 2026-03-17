import os, re, json, base64, uuid, shutil
from pathlib import Path
from typing import List

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── App Setup ────────────────────────────────────────────────────────────────
app = FastAPI(title="OCR Data Extraction Tool")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

MASTER_XLSX = OUTPUT_DIR / "all_extracted_data.xlsx"

HEADERS = [
    "File-Row", "ID Code", "Primary Name", "Location", "Loan Amount",
    "Interest %", "Years", "Rate %", "PV Reduction", "MP Reduction",
    "TI Reduction", "Secondary Name", "Secondary Code"
]

MIME_MAP = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png",  ".webp": "image/webp",
    ".gif": "image/gif"
}

# ── Excel Helpers ────────────────────────────────────────────────────────────
def _style_header(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color="1A1A2E")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    med = Side(style="medium", color="4F46E5")
    cell.border = Border(left=med, right=med, top=med, bottom=med)

def _style_data(cell, row_idx):
    bg = "F0F0FF" if row_idx % 2 == 0 else "FFFFFF"
    thin = Side(style="thin", color="D0D0D0")
    cell.font      = Font(name="Arial", size=9)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="center", wrap_text=True)

COL_WIDTHS = [20, 32, 22, 20, 16, 10, 7, 10, 13, 13, 13, 22, 40]

def create_fresh_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    for col, h in enumerate(HEADERS, 1):
        _style_header(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Extraction Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="1A1A2E")
    ws2["A3"] = "Total Records"; ws2["A3"].font = Font(bold=True, name="Arial")
    ws2["B3"] = 0
    ws2["A5"] = "File";    ws2["A5"].font = Font(bold=True, name="Arial")
    ws2["B5"] = "Records"; ws2["B5"].font = Font(bold=True, name="Arial")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12
    return wb

def append_rows_to_workbook(wb: Workbook, rows: list):
    ws  = wb["Extracted Data"]
    ws2 = wb["Summary"]
    start = ws.max_row + 1
    for ri, row in enumerate(rows, start):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            _style_data(c, ri)
        ws.row_dimensions[ri].height = 22
    total = ws.max_row - 1
    ws2["B3"] = total
    # rebuild file counts
    for r in ws2.iter_rows(min_row=6, max_row=ws2.max_row):
        for c in r: c.value = None
    file_counts: dict = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            fn = str(r[0]).rsplit("-", 1)[0]
            file_counts[fn] = file_counts.get(fn, 0) + 1
    for i, (fn, cnt) in enumerate(file_counts.items(), 6):
        ws2.cell(row=i, column=1, value=fn).font = Font(name="Arial", size=9)
        ws2.cell(row=i, column=2, value=cnt).font  = Font(name="Arial", size=9)

# ── Claude Vision Extraction ─────────────────────────────────────────────────
PROMPT = """Extract ALL loan/mortgage records from this image. Return ONLY a valid JSON array — no markdown, no backticks, no extra text.
Each record must be an object with exactly these keys:
{
  "file_row": "FILENAME-N",
  "id_code": "alphanumeric code exactly as seen",
  "primary_name": "Mr./Ms./Mrs. Full Name",
  "location": "City,State",
  "loan_amount": "numeric digits only — convert written words to number (e.g. Thirty Nine Billion... = 39000000000)",
  "interest_pct": "e.g. 23%",
  "years": "number only",
  "rate_pct": "e.g. 3.64%",
  "pv_reduction": "e.g. 21.87%",
  "mp_reduction": "e.g. 7.76%",
  "ti_reduction": "e.g. 18.87%",
  "secondary_name": "Mr./Ms./Mrs. Full Name",
  "secondary_code": "alphanumeric code exactly as seen"
}
Rules:
- Extract EVERY record visible — miss nothing
- Keep ID/secondary codes exactly as they appear including letters A, E, W, Y, S
- Replace FILENAME with the actual image filename
- Return ONLY the JSON array, nothing else"""

def extract_with_claude(img_path: Path, filename: str) -> list:
    ext  = img_path.suffix.lower()
    mime = MIME_MAP.get(ext, "image/jpeg")
    b64  = base64.standard_b64encode(img_path.read_bytes()).decode()

    client = anthropic.Anthropic()
    resp   = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                {"type": "text",  "text": PROMPT.replace("FILENAME", filename)}
            ]
        }]
    )

    raw     = resp.content[0].text.strip()
    raw     = re.sub(r"```json|```", "", raw).strip()
    records = json.loads(raw)

    rows = []
    for i, r in enumerate(records, 1):
        rows.append([
            r.get("file_row")       or f"{filename}-{i}",
            r.get("id_code",        ""),
            r.get("primary_name",   ""),
            r.get("location",       ""),
            r.get("loan_amount",    ""),
            r.get("interest_pct",   ""),
            r.get("years",          ""),
            r.get("rate_pct",       ""),
            r.get("pv_reduction",   ""),
            r.get("mp_reduction",   ""),
            r.get("ti_reduction",   ""),
            r.get("secondary_name", ""),
            r.get("secondary_code", ""),
        ])
    return rows

# ── Routes ───────────────────────────────────────────────────────────────────
@app.get("/")
def home():
    return {"status": "OCR Tool Running 🚀"}

@app.post("/upload/")
async def upload_files(files: List[UploadFile] = File(...)):
    results = []
    all_new_rows = []

    for upload in files:
        ext = Path(upload.filename).suffix.lower()
        if ext not in MIME_MAP:
            results.append({"filename": upload.filename, "error": "Unsupported file type"})
            continue

        # Save uploaded file
        save_path = UPLOAD_DIR / f"{uuid.uuid4().hex}{ext}"
        with open(save_path, "wb") as f:
            shutil.copyfileobj(upload.file, f)

        try:
            rows = extract_with_claude(save_path, upload.filename)
            all_new_rows.extend(rows)
            results.append({
                "filename": upload.filename,
                "records":  len(rows),
                "rows":     rows
            })
        except Exception as e:
            results.append({"filename": upload.filename, "error": str(e)})
        finally:
            save_path.unlink(missing_ok=True)

    # Append to master Excel
    if all_new_rows:
        if MASTER_XLSX.exists():
            wb = load_workbook(MASTER_XLSX)
        else:
            wb = create_fresh_workbook()
        append_rows_to_workbook(wb, all_new_rows)
        wb.save(MASTER_XLSX)

    total = sum(r.get("records", 0) for r in results)
    return {
        "processed": len(results),
        "total_records": total,
        "results": results,
        "headers": HEADERS,
        "download_url": "/download/" if all_new_rows else None
    }

@app.get("/download/")
def download_excel():
    if not MASTER_XLSX.exists():
        raise HTTPException(status_code=404, detail="No data extracted yet")
    return FileResponse(
        path=str(MASTER_XLSX),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="extracted_data.xlsx"
    )

@app.delete("/reset/")
def reset_data():
    if MASTER_XLSX.exists():
        MASTER_XLSX.unlink()
    return {"message": "Data cleared"}

@app.get("/status/")
def status():
    if not MASTER_XLSX.exists():
        return {"total_records": 0, "file_exists": False}
    wb = load_workbook(MASTER_XLSX, read_only=True)
    ws = wb["Extracted Data"]
    total = ws.max_row - 1
    wb.close()
    return {"total_records": total, "file_exists": True, "download_url": "/download/"}
